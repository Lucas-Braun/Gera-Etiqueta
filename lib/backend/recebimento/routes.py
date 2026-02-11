# -*- coding: utf-8 -*-
from flask import Blueprint, render_template, request, jsonify, send_file
from datetime import datetime, timedelta
import io
import qrcode
import barcode
from barcode.writer import ImageWriter
from PIL import Image
import random
import string
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from lib.backend.recebimento.service import (
    listar_lotes,
    buscar_lote_por_id,
    criar_lote,
    atualizar_lote,
    deletar_lote
)

recebimento_bp = Blueprint('recebimento', __name__, url_prefix='/recebimento')


# ========== PAGINAS PRINCIPAIS ==========

@recebimento_bp.route('/')
@recebimento_bp.route('/home')
def home():
    """Pagina inicial do modulo Recebimento"""
    return render_template('recebimento/home.html')


@recebimento_bp.route('/formulario')
def formulario_lote():
    """Pagina do formulario de lote"""
    return render_template('recebimento/formulario.html')


# ========== APIs ==========

@recebimento_bp.route('/api/lotes')
def api_listar_lotes():
    """API para listar todos os lotes"""
    try:
        lotes = listar_lotes()
        return jsonify({
            'dados': lotes,
            'total_registros': len(lotes)
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@recebimento_bp.route('/api/lotes/<int:lote_id>')
def api_buscar_lote(lote_id):
    """API para buscar lote pelo ID"""
    try:
        lote = buscar_lote_por_id(lote_id)
        if not lote:
            return jsonify({'error': 'Lote nao encontrado'}), 404
        return jsonify(lote)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@recebimento_bp.route('/api/lotes', methods=['POST'])
def api_criar_lote():
    """API para criar novo lote"""
    try:
        dados = request.json

        numero_lote = dados.get('numero_lote', '').strip()
        data_recebimento = dados.get('data_recebimento', '').strip()

        if not numero_lote:
            return jsonify({'error': 'Numero do lote e obrigatorio'}), 400

        # Tratar quantidade
        quantidade = dados.get('quantidade', '').strip()
        quantidade_valor = None
        if quantidade:
            try:
                quantidade_valor = float(quantidade)
            except ValueError:
                return jsonify({'error': 'Quantidade deve ser um numero valido'}), 400

        lote, erro = criar_lote(
            numero_lote=numero_lote,
            id_item=dados.get('id_item', '').strip() or None,
            data_recebimento=data_recebimento or None,
            data_fabricacao=dados.get('data_fabricacao', '').strip() or None,
            data_validade=dados.get('data_validade', '').strip() or None,
            quantidade=quantidade_valor,
            numero_nota_fiscal=dados.get('numero_nota_fiscal', '').strip() or None,
            observacao=dados.get('observacao', '').strip() or None
        )

        if erro:
            return jsonify({'error': erro}), 400

        return jsonify({
            'success': True,
            'message': 'Lote salvo com sucesso!',
            'lote': lote
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@recebimento_bp.route('/api/lotes/<int:lote_id>', methods=['PUT'])
def api_atualizar_lote(lote_id):
    """API para atualizar lote existente"""
    try:
        dados = request.json

        numero_lote = dados.get('numero_lote', '').strip()
        data_recebimento = dados.get('data_recebimento', '').strip()

        if not numero_lote:
            return jsonify({'error': 'Numero do lote e obrigatorio'}), 400

        # Tratar quantidade
        quantidade = dados.get('quantidade', '').strip()
        quantidade_valor = None
        if quantidade:
            try:
                quantidade_valor = float(quantidade)
            except ValueError:
                return jsonify({'error': 'Quantidade deve ser um numero valido'}), 400

        lote, erro = atualizar_lote(
            lote_id=lote_id,
            numero_lote=numero_lote,
            id_item=dados.get('id_item', '').strip() or None,
            data_recebimento=data_recebimento or None,
            data_fabricacao=dados.get('data_fabricacao', '').strip() or None,
            data_validade=dados.get('data_validade', '').strip() or None,
            quantidade=quantidade_valor,
            numero_nota_fiscal=dados.get('numero_nota_fiscal', '').strip() or None,
            observacao=dados.get('observacao', '').strip() or None
        )

        if erro:
            return jsonify({'error': erro}), 400

        return jsonify({
            'success': True,
            'message': 'Lote atualizado com sucesso!',
            'lote': lote
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@recebimento_bp.route('/api/lotes/<int:lote_id>', methods=['DELETE'])
def api_deletar_lote(lote_id):
    """API para deletar lote - BLOQUEADO"""
    return jsonify({'error': 'Exclusao de lotes esta desabilitada'}), 403


@recebimento_bp.route('/api/gerar-etiqueta', methods=['POST'])
def api_gerar_etiqueta():
    """API para gerar etiqueta em PDF com QR Code"""
    try:
        from reportlab.pdfgen import canvas
        from reportlab.lib.units import mm
        from datetime import timedelta

        dados = request.json

        # Configuracoes da etiqueta
        largura_mm = float(dados.get('largura', 100))
        altura_mm = float(dados.get('altura', 75))
        tamanho_fonte_base = int(dados.get('tamanho_fonte', 8))
        quantidade_etiquetas = int(dados.get('quantidade_etiquetas', 1))
        quantidade_etiquetas = max(1, quantidade_etiquetas)

        # Converter mm para pontos
        largura = largura_mm * mm
        altura = altura_mm * mm

        # Criar buffer para o PDF
        buffer = io.BytesIO()

        # Criar o canvas PDF
        c = canvas.Canvas(buffer, pagesize=(largura, altura))

        # Sistema adaptativo de fonte
        if largura <= 60 * mm or altura <= 50 * mm:
            x_margin = 1 * mm
            y_margin = 1 * mm
            tamanho_fonte = max(4, tamanho_fonte_base - 2)
            titulo_fonte = tamanho_fonte + 1
            lote_fonte = tamanho_fonte + 1
            rodape_fonte = tamanho_fonte - 1
            line_height = (tamanho_fonte + 2) * 1.2
            section_space = 0
        elif largura <= 90 * mm or altura <= 70 * mm:
            x_margin = 1.5 * mm
            y_margin = 1.5 * mm
            tamanho_fonte = max(5, tamanho_fonte_base - 1)
            titulo_fonte = tamanho_fonte + 2
            lote_fonte = tamanho_fonte + 2
            rodape_fonte = tamanho_fonte - 1
            line_height = (tamanho_fonte + 2) * 1.2
            section_space = 0
        else:
            x_margin = 2 * mm
            y_margin = 2 * mm
            tamanho_fonte = tamanho_fonte_base
            titulo_fonte = tamanho_fonte + 3
            lote_fonte = tamanho_fonte + 2
            rodape_fonte = tamanho_fonte - 2
            line_height = (tamanho_fonte + 2) * 1.2
            section_space = 0

        area_util_largura = largura - (2 * x_margin)
        area_util_altura = altura - (2 * y_margin)
        y_start = altura - y_margin - lote_fonte

        # Gerar codigo para QR Code - somente ID do Item
        numero_lote = dados.get('numero_lote', '').strip()
        id_item_qr = dados.get('id_item', '').strip()

        # QR Code contem apenas o ID do Item
        codigo_qr = id_item_qr if id_item_qr else 'SEM_ID'
        # Garantir que e ASCII puro
        codigo_qr = codigo_qr.encode('ascii', 'replace').decode('ascii')

        # Gerar QR Code
        qr = qrcode.QRCode(
            version=1,
            error_correction=qrcode.constants.ERROR_CORRECT_M,
            box_size=4,
            border=2,
        )
        qr.add_data(codigo_qr)
        qr.make(fit=True)
        qr_img = qr.make_image(fill_color="black", back_color="white")

        # Gerar codigo de barras
        barcode_data = codigo_qr
        barcode_class = barcode.get_barcode_class('code128')
        barcode_obj = barcode_class(barcode_data, writer=ImageWriter())
        barcode_buffer = io.BytesIO()
        barcode_obj.write(barcode_buffer, options={
            'write_text': False,
            'module_height': 8,
            'module_width': 0.25,
            'quiet_zone': 1
        })
        barcode_buffer.seek(0)
        barcode_img = Image.open(barcode_buffer)

        # Loop para gerar multiplas etiquetas
        for etiqueta_num in range(quantidade_etiquetas):
            texto_x = x_margin
            agora_brasilia = datetime.now() - timedelta(hours=3)

            # === ZONA INFERIOR: rodape + codigo de barras (fixo no fundo) ===
            barcode_altura = 8 * mm
            rodape_y = y_margin
            barcode_y = y_margin + rodape_fonte + (1 * mm)
            zona_inferior_topo = barcode_y + barcode_altura + (1 * mm)

            # === ZONA SUPERIOR: texto + QR (do topo ate a zona inferior) ===
            zona_superior_altura = y_start + lote_fonte - zona_inferior_topo

            # QR adaptativo - ocupa a altura disponivel na zona superior
            qr_size = min(largura * 0.33, zona_superior_altura)
            qr_size = max(12 * mm, qr_size)
            qr_area_x = largura - x_margin - qr_size
            qr_area_y = y_start + lote_fonte - qr_size  # alinhado ao topo

            texto_limite_x = qr_area_x - (2 * mm)

            # Contar linhas de texto para distribuir uniformemente
            linhas = []
            if numero_lote:
                linhas.append(('bold_lote', f"LOTE: {numero_lote}"))
            id_item = dados.get('id_item', '')
            if id_item:
                linhas.append(('bold', f"ID Item: {id_item}"))
            descricao_item = dados.get('descricao_item', '')
            if descricao_item:
                texto_max_largura = texto_limite_x - texto_x
                desc_max = max(20, int(texto_max_largura / (tamanho_fonte * 0.5)))
                desc_texto = descricao_item[:desc_max] + "..." if len(descricao_item) > desc_max else descricao_item
                linhas.append(('normal', desc_texto))
            data_recebimento = dados.get('data_recebimento', '')
            if data_recebimento:
                linhas.append(('normal', f"Recebimento: {data_recebimento}"))
            data_fabricacao = dados.get('data_fabricacao', '')
            if data_fabricacao:
                linhas.append(('normal', f"Fabricacao: {data_fabricacao}"))
            data_validade = dados.get('data_validade', '')
            if data_validade:
                linhas.append(('bold', f"Validade: {data_validade}"))
            numero_nota_fiscal = dados.get('numero_nota_fiscal', '')
            if numero_nota_fiscal:
                linhas.append(('normal', f"Nota Fiscal: {numero_nota_fiscal}"))
            quantidade_prod = dados.get('quantidade', '')
            if quantidade_prod:
                linhas.append(('normal', f"Quantidade: {quantidade_prod}"))
            observacao = dados.get('observacao', '')
            if observacao:
                texto_max_largura = texto_limite_x - texto_x
                obs_max = max(20, int(texto_max_largura / (tamanho_fonte * 0.5)))
                obs_texto = observacao[:obs_max] + "..." if len(observacao) > obs_max else observacao
                linhas.append(('small', f"Obs: {obs_texto}"))

            # Desenhar linhas com espacamento compacto
            y = y_start
            for tipo, texto in linhas:
                if tipo == 'bold_lote':
                    c.setFont("Helvetica-Bold", lote_fonte)
                elif tipo == 'bold':
                    c.setFont("Helvetica-Bold", tamanho_fonte)
                elif tipo == 'small':
                    c.setFont("Helvetica", max(3, tamanho_fonte - 1))
                else:
                    c.setFont("Helvetica", tamanho_fonte)
                c.drawString(texto_x, y, texto)
                y -= line_height

            # QR CODE (lado direito)
            c.drawInlineImage(qr_img, qr_area_x, qr_area_y, qr_size, qr_size)

            # CODIGO DE BARRAS (abaixo do conteudo - texto ou QR, o mais baixo)
            barcode_largura_img = largura - (2 * x_margin)
            conteudo_bottom = min(y, qr_area_y)
            barcode_y_calc = conteudo_bottom - barcode_altura - (1 * mm)
            barcode_y_calc = max(barcode_y_calc, y_margin + rodape_fonte + (1 * mm))
            c.drawInlineImage(barcode_img, x_margin, barcode_y_calc, barcode_largura_img, barcode_altura)

            # RODAPE
            c.setFont("Helvetica", rodape_fonte)
            rodape_texto = f"ID: {codigo_qr}"
            rodape_largura_txt = c.stringWidth(rodape_texto, "Helvetica", rodape_fonte)
            rodape_x = (largura - rodape_largura_txt) / 2
            rodape_y_calc = barcode_y_calc - rodape_fonte - (0.5 * mm)
            rodape_y_calc = max(rodape_y_calc, y_margin)
            c.drawString(rodape_x, rodape_y_calc, rodape_texto)

            # Criar nova pagina se nao for a ultima etiqueta
            if etiqueta_num < quantidade_etiquetas - 1:
                c.showPage()

        # Finalizar
        c.save()
        buffer.seek(0)

        nome_arquivo = f'etiqueta_{numero_lote or "sem_lote"}_{agora_brasilia.strftime("%Y%m%d_%H%M%S")}.pdf'

        return send_file(
            buffer,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=nome_arquivo
        )

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@recebimento_bp.route('/api/exportar-xls')
def exportar_xls():
    """Exportar todos os lotes para Excel"""
    try:
        lotes = listar_lotes()

        wb = Workbook()
        ws = wb.active
        ws.title = "Lotes Recebimento"

        # Estilo do cabecalho
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="28A745", end_color="28A745", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        colunas = ["ID", "Número do Lote", "ID Item", "Data Recebimento", "Data Fabricação",
                    "Data Validade", "Quantidade", "Nota Fiscal", "Observação"]

        for col_num, titulo in enumerate(colunas, 1):
            cell = ws.cell(row=1, column=col_num, value=titulo)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Dados
        for row_num, lote in enumerate(lotes, 2):
            valores = [
                lote.get('id', ''),
                lote.get('numero_lote', ''),
                lote.get('id_item', ''),
                lote.get('data_recebimento', ''),
                lote.get('data_fabricacao', ''),
                lote.get('data_validade', ''),
                lote.get('quantidade', ''),
                lote.get('numero_nota_fiscal', ''),
                lote.get('observacao', '')
            ]
            for col_num, valor in enumerate(valores, 1):
                cell = ws.cell(row=row_num, column=col_num, value=valor)
                cell.border = thin_border

        # Ajustar largura das colunas
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_length + 4, 40)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        agora = (datetime.now() - timedelta(hours=3)).strftime("%Y%m%d_%H%M%S")
        nome_arquivo = f'lotes_recebimento_{agora}.xlsx'

        return send_file(
            buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=nome_arquivo
        )

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500
